from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from itertools import combinations

class Grupo:
    def __init__(self, clave, grupo, profesor, tipo, horario, dias):
        self.clave = clave
        self.grupo = grupo
        self.profesor = profesor
        self.tipo = tipo
        self.horario = horario  # [hora_inicio, hora_fin]
        self.dias = dias  # Lista de enteros representando los días de la semana (0-6)

    def __repr__(self):
        return f"Grupo(clave={self.clave}, grupo={self.grupo}, profesor={self.profesor}, tipo={self.tipo}, horario={self.horario}, dias={self.dias})"

def formatearObjetos(asignaturas):
    for asignatura in asignaturas:
        asignatura.profesor = asignatura.profesor.split('(', 1)[0]
        hora_inicio, hora_fin = asignatura.horario.split(' a ')
        asignatura.horario = [hora_inicio, hora_fin]
        dias_numerico = []
        for dia in asignatura.dias.split(','):  # Directly iterate over the split result
            dia = dia.strip()  # Strip spaces from each day
            if dia == 'Lun':
                dias_numerico.append(1)
            elif dia == 'Mar':
                dias_numerico.append(2)
            elif dia == 'Mie':
                dias_numerico.append(3)
            elif dia == 'Jue':
                dias_numerico.append(4)
            elif dia == 'Vie':
                dias_numerico.append(5)
            elif dia == 'Sab':
                dias_numerico.append(6)
        asignatura.dias = dias_numerico
        

def es_vespertino(grupo):
    return any(hora >= '15:00' for hora in grupo.horario)

def separar_grupos_por_horario(grupos):
    matutinos = [grupo for grupo in grupos if not es_vespertino(grupo)]
    vespertinos = [grupo for grupo in grupos if es_vespertino(grupo)]
    return matutinos, vespertinos

def horarios_no_se_traslapen(horario1, dias1, horario2, dias2):
    dias_comunes = set(dias1) & set(dias2)
    if not dias_comunes:
        return True

    inicio1, fin1 = [int(h.replace(':', '')) for h in horario1]
    inicio2, fin2 = [int(h.replace(':', '')) for h in horario2]
    return fin1 <= inicio2 or fin2 <= inicio1


def generar_combinaciones(grupos, claves_asignaturas):
    posibles_horarios = []

    for combinacion in combinations(grupos, len(claves_asignaturas)):
        if len(set(grupo.clave for grupo in combinacion)) == len(claves_asignaturas):
            valido = True
            for i, grupo1 in enumerate(combinacion):
                for grupo2 in combinacion[i+1:]:
                    if not horarios_no_se_traslapen(grupo1.horario, grupo1.dias, grupo2.horario, grupo2.dias):
                        valido = False
                        break
                if not valido:
                    break
            if valido:
                posibles_horarios.append(combinacion)
    
    return posibles_horarios

def crear_horario_excel(horarios, archivo_salida):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario Semanal"

    # Cabeceras de días de la semana
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    horas = [f"{h:02d}:{m:02d}" for h in range(7, 24) for m in (0, 30)]
    
    fila_inicio = 1
    for i, horario in enumerate(horarios):
        # Escribir cabeceras
        ws.cell(row=fila_inicio, column=1, value='Hora')
        for col, dia in enumerate(dias_semana, start=2):
            ws.cell(row=fila_inicio, column=col, value=dia)
        ws.cell(row=fila_inicio, column=len(dias_semana) + 2, value='Datos del Grupo')

        # Escribir rango de horas
        for row, hora in enumerate(horas, start=fila_inicio + 1):
            ws.cell(row=row, column=1, value=hora)

        # Ubicar los horarios en las celdas correspondientes
        for grupo in horario:
            inicio, fin = grupo.horario
            datos_grupo = f"{grupo.clave} - {grupo.grupo}\n{grupo.profesor}"

            inicio_idx = horas.index(inicio) + fila_inicio + 1
            fin_idx = horas.index(fin) + fila_inicio + 1
            for dia in grupo.dias:
                for fila in range(inicio_idx, fin_idx):
                    ws.cell(row=fila, column=dia + 1, value=f"{grupo.clave}-{grupo.grupo}")
                    ws.cell(row=fila, column=dia + 1).alignment = Alignment(horizontal='center', vertical='center')
                
                # Añadir datos del grupo al lado derecho
                ws.cell(row=inicio_idx, column=len(dias_semana) + 2, value=datos_grupo)
                ws.cell(row=inicio_idx, column=len(dias_semana) + 2).alignment = Alignment(horizontal='left', vertical='top')

        # Incrementar la fila de inicio para la próxima tabla con un espaciado de 5 filas
        fila_inicio += len(horas) + 6

    wb.save(archivo_salida)

# Configura el servicio de EdgeDriver (suponiendo que msedgedriver esté en el PATH)

# Abre la página web

asignaturas = []  # Lista para almacenar objetos Asignatura

def obtenerDatos(arreglo_materias):
    grupos = [
    Grupo('1644', '4', 'ING. LUCIRALIA HERNANDEZ HERNANDEZ', 'T', ['15:00','17:00'], [1, 3, 5]),
    Grupo('1644', '6', 'M.C. DAVID RICARDO RUIZ REYES', 'T', ['17:00','19:00'], [2, 3, 4]),
    Grupo('1562', '1', 'ING. VICTOR MANUEL SANCHEZ ESQUIVEL', 'T', ['16:00','17:30'], [2, 4]),
    Grupo('1562', '2', 'M.I. PATRICIA HONG CIRION', 'T', ['18:00','19:30'], [2, 4]),
    Grupo('1562', '3', 'ING. IVAN MARTINEZ PEREZ', 'T', ['20:30','22:00'], [2, 4]),
    ]

    return grupos


def main():
    # Base de Datos, Circuitos Electricos, Finanzas, Inteligencia Artificial, Economia
    # arreglo_materias = [1644, 1562, 1537, 406, 1413]
    arreglo_materias = [1644,1562]
    grupos = obtenerDatos(arreglo_materias)
    #Le damos formato de arreglo al horario y dias
    # formatearObjetos(grupos)

    matutinos, vespertinos = separar_grupos_por_horario(grupos)

    posibles_horarios_vespertinos = generar_combinaciones(vespertinos, arreglo_materias)
    posibles_horarios_matutinos = generar_combinaciones(matutinos, arreglo_materias)

    # Guardar horarios en archivos Excel
    if posibles_horarios_vespertinos:
        opciones = []
        for horario in posibles_horarios_vespertinos:
            if len(horario) == len(arreglo_materias):
                opciones.append(horario)
        
        crear_horario_excel(opciones, 'HorariosVespertinos.xlsx')
        print("Horario guardado en 'HorariosVespertinos.xlsx'")
    else:
        print("No hay horarios vespertinos disponibles")

    if posibles_horarios_matutinos:
        opciones = []
        for horario in posibles_horarios_matutinos:
            if len(horario) == len(arreglo_materias):
                opciones.append(horario)
        
        crear_horario_excel(opciones, 'HorariosMatutinos.xlsx')
        print("Horario guardado en 'HorariosMatutinos.xlsx'")
    else:
        print("No hay horarios matutinos disponibles")
    
main()