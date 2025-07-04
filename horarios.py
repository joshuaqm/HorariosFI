'''
AUTOR: FRANCISCO JOSHUA QUINTERO MONTERO

ARMADOR DE HORARIOS DE LA FACULTAD DE INGENIERÍA UNAM
Este script permite obtener los horarios de las asignaturas de la Facultad de Ingeniería de la UNAM
y generar horarios personalizados para un conjunto de asignaturas específicas.

Requisitos:
- Python 3.6 o superior
- Selenium (pip install selenium)
- BeautifulSoup (pip install beautifulsoup4)
- Openpyxl (pip install openpyxl)
- WebDriver Manager (pip install webdriver-manager)
- itertools (incluido en Python por defecto)

'''
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment
from itertools import combinations

class Grupo:
    def __init__(self, clave, grupo, profesor, tipo, horario, dias):
        self.clave = clave
        self.grupo = grupo
        self.profesor = profesor
        self.tipo = tipo
        self.horario = horario  # [hora_inicio, hora_fin]
        self.dias = dias  # Lista de enteros representando los días de la semana (0-6)

    def __repr__(self):
        return f"Grupo(clave={self.clave}, grupo={self.grupo}, profesor={self.profesor}, tipo={self.tipo}, horario={self.horario}, dias={self.dias})"

def formatearObjetos(asignaturas):
    for asignatura in asignaturas:
        asignatura.profesor = asignatura.profesor.split('(', 1)[0]
        hora_inicio, hora_fin = asignatura.horario.split(' a ')
        asignatura.horario = [hora_inicio, hora_fin]
        dias_numerico = []
        for dia in asignatura.dias.split(','):  # Directly iterate over the split result
            dia = dia.strip()  # Strip spaces from each day
            if dia == 'Lun':
                dias_numerico.append(1)
            elif dia == 'Mar':
                dias_numerico.append(2)
            elif dia == 'Mie':
                dias_numerico.append(3)
            elif dia == 'Jue':
                dias_numerico.append(4)
            elif dia == 'Vie':
                dias_numerico.append(5)
            elif dia == 'Sab':
                dias_numerico.append(6)
        asignatura.dias = dias_numerico
        

def es_vespertino(grupo):
    return any(hora >= '15:00' for hora in grupo.horario)

def separar_grupos_por_horario(grupos):

    matutinos = [grupo for grupo in grupos if not es_vespertino(grupo)]
    vespertinos = [grupo for grupo in grupos if es_vespertino(grupo)]
    return matutinos, vespertinos

def horarios_no_se_traslapen(horario1, dias1, horario2, dias2):
    dias_comunes = set(dias1) & set(dias2)
    if not dias_comunes:
        return True

    inicio1, fin1 = [int(h.replace(':', '')) for h in horario1]
    inicio2, fin2 = [int(h.replace(':', '')) for h in horario2]
    return fin1 <= inicio2 or fin2 <= inicio1


def generar_combinaciones(grupos, claves_asignaturas):
    posibles_horarios = []

    for combinacion in combinations(grupos, len(claves_asignaturas)):
        if len(set(grupo.clave for grupo in combinacion)) == len(claves_asignaturas):
            valido = True
            for i, grupo1 in enumerate(combinacion):
                for grupo2 in combinacion[i+1:]:
                    if not horarios_no_se_traslapen(grupo1.horario, grupo1.dias, grupo2.horario, grupo2.dias):
                        valido = False
                        break
                if not valido:
                    break
            if valido:
                posibles_horarios.append(combinacion)
    
    return posibles_horarios

def crear_horario_excel(horarios, archivo_salida):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Horario Semanal"

        # Cabeceras de días de la semana
        dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
        horas = [f"{h:02d}:{m:02d}" for h in range(7, 24) for m in (0, 30)]
        
        fila_inicio = 1
        for i, horario in enumerate(horarios):
            # Escribir cabeceras
            ws.cell(row=fila_inicio, column=1, value='Hora')
            for col, dia in enumerate(dias_semana, start=2):
                ws.cell(row=fila_inicio, column=col, value=dia)
            ws.cell(row=fila_inicio, column=len(dias_semana) + 2, value='Datos del Grupo')

            # Escribir rango de horas
            for row, hora in enumerate(horas, start=fila_inicio + 1):
                ws.cell(row=row, column=1, value=hora)

            # Ubicar los horarios en las celdas correspondientes
            for grupo in horario:
                inicio, fin = grupo.horario
                datos_grupo = f"{grupo.clave} - {grupo.grupo} - {grupo.profesor}"

                inicio_idx = horas.index(inicio) + fila_inicio + 1
                fin_idx = horas.index(fin) + fila_inicio + 1

                for dia in grupo.dias:
                    for fila in range(inicio_idx, fin_idx):
                        # Escribir el grupo en el horario correspondiente
                        if ws.cell(row=fila, column=dia + 1).value:
                            ws.cell(row=fila, column=dia + 1).value += f", {grupo.clave}-{grupo.grupo}"
                        else:
                            ws.cell(row=fila, column=dia + 1, value=f"{grupo.clave}-{grupo.grupo}")
                        ws.cell(row=fila, column=dia + 1).alignment = Alignment(horizontal='center', vertical='center')

                # Añadir datos del grupo al lado derecho
                grupo_fila = fila_inicio + 1 + horario.index(grupo) * 2  # Incrementar la fila por cada grupo
                ws.cell(row=grupo_fila, column=len(dias_semana) + 2, value=datos_grupo)
                ws.cell(row=grupo_fila, column=len(dias_semana) + 2).alignment = Alignment(horizontal='left', vertical='top')

            # Incrementar la fila de inicio para la próxima tabla con un espaciado de 5 filas
            fila_inicio += len(horas) + 6

        wb.save(archivo_salida)
        return True
    except Exception as e:
        print(f"Error: {e}")
        return False


asignaturas = []  # Lista para almacenar objetos Asignatura

def obtenerDatos(arreglo_materias):
    # Configura el servicio de EdgeDriver con opciones
    from selenium.webdriver.edge.service import Service
    from selenium.webdriver.edge.options import Options
    
    # Configurar opciones del navegador
    edge_options = Options()
    edge_options.add_experimental_option('excludeSwitches', ['enable-logging'])
    edge_options.add_experimental_option('useAutomationExtension', False)
    edge_options.add_argument('--disable-blink-features=AutomationControlled')
    edge_options.add_argument('--disable-extensions')
    # edge_options.add_argument('--headless')  # Opcional: ejecutar sin interfaz gráfica
    
    # Intentar diferentes formas de inicializar el driver
    driver = None
    try:
        # Método 1: Usando Service (más moderno)
        try:
            service = Service()
            driver = webdriver.Edge(service=service, options=edge_options)
        except Exception:
            # Método 2: Usando webdriver-manager (automático)
            try:
                from webdriver_manager.microsoft import EdgeChromiumDriverManager
                service = Service(EdgeChromiumDriverManager().install())
                driver = webdriver.Edge(service=service, options=edge_options)
            except ImportError:
                print("webdriver-manager no está instalado. Instálalo con: pip install webdriver-manager")
                return []
        
        # Abre la página web
        driver.get("https://www.ssa.ingenieria.unam.mx/horarios.html")
        
        # Espera adicional para que la página se cargue completamente
        import time
        time.sleep(3)
        
        # Espera a que el formulario esté presente y visible
        form = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "hrsFormHorarioAsignatura"))
        )
        
        # Espera a que los elementos del formulario estén presentes e interactuables
        clave_input = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "clave"))
        )
        buscar_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "buscar"))
        )

        prioridad = 0
        # Llena los campos del formulario
        for clave in arreglo_materias:
            prioridad += 1
            
            # Asegurarse de que el campo esté limpio
            clave_input.clear()
            time.sleep(1)
            
            # Enviar la clave
            clave_input.send_keys(str(clave))
            time.sleep(1)
            
            # Hacer clic en el botón de búsqueda
            buscar_button.click()
            
            # Esperar a que los resultados se carguen
            time.sleep(3)
            
            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            # Encuentra la tabla por su clase
            table = soup.find('table', {'class': 'table table-horarios-custom'})
            if table:
                # Procesa la tabla y crea un DataFrame
                rows = table.find_all('tr')
                for row in rows:
                    cols = row.find_all('td')
                    cols = [col.text.strip() for col in cols]
                    # Verifica si la fila tiene datos válidos
                    if cols and len(cols) == 7:  # Asumiendo que cada fila tiene 7 columnas
                        # Crea un objeto Asignatura y agrégalo a la lista
                        profesor = cols[2].split('(', 1)[0]
                        asignatura = Grupo(cols[0], cols[1], cols[2], cols[3], cols[4], cols[5])
                        asignaturas.append(asignatura)
                print(f"Datos obtenidos para la clave de asignatura: {clave}")
            else:
                print(f"No se encontró la clave de asignatura: {clave} en la página de horarios.")
                
            # Limpiar el campo para la siguiente búsqueda
            clave_input.clear()
            time.sleep(1)
        return asignaturas
    except Exception as e:
        # En caso de error, captura la excepción y muestra un mensaje
        print("Error al cargar la página:", e)
        return []

    finally:
        # Cierra el navegador al finalizar
        if driver:
            driver.quit()

def filtrar_grupos_por_dias_y_horas(grupos, dias_deseados, horario_deseado):
    grupos_filtrados = []
    entrada_deseada, salida_deseada = [int(h.replace(':', '')) for h in horario_deseado]
    
    for grupo in grupos:
        horario_inicio, horario_fin = [int(h.replace(':', '')) for h in grupo.horario]
        
        # Verificar si los días del grupo están dentro de los días deseados
        if all(dia in dias_deseados for dia in grupo.dias):
            # Verificar si el horario del grupo está dentro del horario deseado
            if entrada_deseada <= horario_inicio and horario_fin <= salida_deseada:
                grupos_filtrados.append(grupo)
    
    return grupos_filtrados

def main():
    # Base de Datos, Lab BD, Circuitos Electricos, Lab Circuitos, Finanzas, Inteligencia Artificial, Economia
    # arreglo_materias = [1644, 6644, 1562, 6562, 1537, 406, 1413]
    # arreglo_materias = [1562, 6562, 1537, 1535, 406, 434, 1686, 6686, 1413]
    # Arreglo de claves de asignaturas
    arreglo_materias = [1598, 1672, 2080, 2929]
    # Arreglo de dias deseados para asistir a clases (1=Lunes, 2=Martes, 3=Miercoles, 4=Jueves, 5=Viernes, 6=Sabado)
    arreglo_dias = [1, 3, 4] 
    # Arreglo de horarios deseados para asistir a clases (hora de entrada, hora de salida)
    horario_deseado = ['09:00', '21:00'] 
    
    grupos = obtenerDatos(arreglo_materias)
    print(f"Total de grupos obtenidos: {len(grupos)}")
    
    # Le damos formato de arreglo al horario y dias
    formatearObjetos(grupos)
    
    # Filtrar grupos por días y horas deseadas
    grupos_filtrados = filtrar_grupos_por_dias_y_horas(grupos, arreglo_dias, horario_deseado)
    print(f"Grupos filtrados: {len(grupos_filtrados)}")
    
    # Mostrar información de los grupos filtrados
    for grupo in grupos_filtrados:
        print(f"Clave: {grupo.clave}, Grupo: {grupo.grupo}, Horario: {grupo.horario}, Días: {grupo.dias}")

    posibles_horarios = generar_combinaciones(grupos_filtrados, arreglo_materias)
    if posibles_horarios:
        opciones = []
        for horario in posibles_horarios:
            if len(horario) == len(arreglo_materias):
                opciones.append(horario)
        
        success = crear_horario_excel(opciones, 'Horarios.xlsx')
        if success:
            print("Opciones generadas:", len(opciones))
            print("Horario guardado en 'Horarios.xlsx'")
        else:
            print("Error al guardar el archivo")
    else:
        print('No hay opciones disponibles')

main()
